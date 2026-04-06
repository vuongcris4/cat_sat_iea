from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
import json
import logging
from logging.handlers import RotatingFileHandler
import os
from .forms import OptimizationForm
from .optimization_logic import SteelCuttingOptimizer, SolverTimer
import asyncio
from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync
import sys
import traceback
from datetime import datetime

# Configure logging với file storage
LOG_DIR = '/app/logs'
os.makedirs(LOG_DIR, exist_ok=True)

logger = logging.getLogger('cat_sat')
logger.setLevel(logging.INFO)

# Console handler
console_handler = logging.StreamHandler()
console_handler.setFormatter(logging.Formatter('%(asctime)s [CAT_SAT] %(levelname)s: %(message)s'))
logger.addHandler(console_handler)

# File handler với rotation (10MB per file, giữ 5 files)
file_handler = RotatingFileHandler(
    os.path.join(LOG_DIR, 'cat_sat.log'),
    maxBytes=10*1024*1024,  # 10MB
    backupCount=5,
    encoding='utf-8'
)
file_handler.setFormatter(logging.Formatter('%(asctime)s [CAT_SAT] %(levelname)s: %(message)s'))
logger.addHandler(file_handler)

class TeeStream:
    def __init__(self, websocket_room):
        self.websocket_room = websocket_room
        self.channel_layer = get_channel_layer()

    def write(self, message):
        if message.strip():
            async_to_sync(self.channel_layer.group_send)(
                f"{self.websocket_room}",
                {"type": "chat.message", "message": message}
            )

    def flush(self):
        pass

@login_required
def index(request):
    form = OptimizationForm()
    return render(request, 'cat_sat/index.html', {'form': form})

@login_required
def optimize(request):
    if request.method == 'POST':
        original_stdout = sys.stdout
        try:
            sys.stdout = TeeStream("log_gurobi_solver_cat_sat")
            
            # === LOGGING USER REQUEST ===
            logger.info("="*60)
            logger.info(f"NEW OPTIMIZATION REQUEST at {datetime.now().isoformat()}")
            logger.info(f"User: {request.user.username}")
            logger.info("="*60)
            data = json.loads(request.body)

            time_limit_minutes = data.get('time_limit_minutes', 2)
            time_limit_seconds = int(time_limit_minutes) * 60
            timer = SolverTimer(time_limit_seconds)

            try:
                # === THAY ĐỔI LOGIC LẤY DỮ LIỆU ===
                pieces_data = data.get('pieces_data')
                if not pieces_data:
                    print("❌ Không có dữ liệu đoạn cắt được cung cấp.<br>")
                    return JsonResponse({'status': 'error', 'message': 'Không có dữ liệu đoạn cắt.'}, status=400)

                # Lọc các hàng rỗng từ handsontable (hàng phải có cả 3 cột)
                valid_pieces = [row for row in pieces_data if row and row[0] is not None and row[1] is not None and row[2] is not None]
                
                if not valid_pieces:
                    print("❌ Dữ liệu đoạn cắt rỗng.<br>")
                    return JsonResponse({'status': 'error', 'message': 'Dữ liệu đoạn cắt rỗng.'}, status=400)

                # Tách dữ liệu từ bảng (3 cột)
                piece_names = [str(item[0]) for item in valid_pieces]
                segment_sizes = [float(item[1]) for item in valid_pieces]
                demands = [int(item[2]) for item in valid_pieces]

                # Lấy factors (dạng string) và chuyển thành list int
                factors_str = data.get('factors', "")
                factors = [int(f) for f in factors_str.replace('.', ' ').replace(',', ' ').split() if f.isdigit()]
                if not factors: # Nếu rỗng, mặc định là [1]
                    factors = [1] 
                # === KẾT THÚC THAY ĐỔI ===

                # === LOGGING INPUT PARAMETERS ===
                logger.info(f"Stock Length: {data['length']}mm")
                logger.info(f"Te Dau Sat: {data['te_dau_sat']}mm")
                logger.info(f"Blade Width: {data['blade_width']}mm")
                logger.info(f"Max Manual Cuts: {data['max_manual_cuts']}")
                logger.info(f"Max Stock Over: {data['max_stock_over']}")
                logger.info(f"Hao Hut Percent: {data.get('hao_hut_percent', 1)}%")
                logger.info(f"Time Limit: {time_limit_seconds}s")
                logger.info(f"Factors: {factors}")
                logger.info(f"Number of piece types: {len(piece_names)}")
                for i, (name, size, demand) in enumerate(zip(piece_names, segment_sizes, demands)):
                    logger.info(f"  [{i+1}] {name}: {size}mm x {demand} pcs")
                logger.info("-"*60)

                optimizer = SteelCuttingOptimizer(
                    length=int(data['length']),
                    te_dau_sat=int(data['te_dau_sat']),
                    piece_names=piece_names,       # Thêm Tên sắt
                    segment_sizes=segment_sizes, # Truyền list đã tách
                    demands=demands,           # Truyền list đã tách
                    blade_width=float(data['blade_width']),
                    factors=factors,           # Truyền list int
                    max_manual_cuts=int(data['max_manual_cuts']),
                    max_stock_over=int(data['max_stock_over']),
                    hao_hut_percent=float(data.get('hao_hut_percent', 1)),
                    time_limit_seconds=time_limit_seconds
                )

                logger.info("Starting Phase 1: Pattern Generation...")
                solutions = optimizer.optimize_cutting()
                logger.info(f"Phase 1 complete. Found {len(solutions) if solutions else 0} patterns.")
                
                logger.info("Starting Phase 2: Distribution Optimization...")
                timer.start()
                distribution = optimizer.optimize_distribution()
                logger.info("Phase 2 complete.")

                # === LOG OPTIMIZATION RESULTS ===
                logger.info("="*60)
                logger.info("OPTIMIZATION RESULTS")
                logger.info(f"  Total Patterns Found: {len(solutions) if solutions else 0}")
                logger.info(f"  Distribution Shape: {distribution.shape if hasattr(distribution, 'shape') else len(distribution)}")
                if hasattr(distribution, 'sum'):
                    logger.info(f"  Total Bars Used: {int(distribution.sum())}")
                logger.info("="*60)
                
                return JsonResponse({
                    "status": "success",
                    "solutions": solutions,
                    "distribution": distribution.tolist()
                }, status=200)
            
            finally:
                timer.stop()
                if timer.is_alive():
                    timer.join()
        
        except Exception as e:
            # Lấy chi tiết toàn bộ traceback_message
            traceback_message = traceback.format_exc() 
            
            # In ra WebSocket (vì stdout đang bị TeeStream bắt)
            print("!CLEAR!") # Xóa log cũ
            print("❌ ĐÃ XẢY RA LỖI NGHIÊM TRỌNG (GĐ 2) ❌<br>")
            print("<pre style='color:red; font-family: monospace;'>")
            print(traceback_message)
            print("</pre>")

            return JsonResponse({"status": "error", "message": str(e)}, status=500)
        
        finally:
            sys.stdout = original_stdout # <-- Sửa lại thành __stdout__

    return JsonResponse({"status": "error", "message": "Invalid request method"}, status=400)


@login_required
def export_excel_phase1(request):
    """Xuất kết quả GĐ 1 (danh sách patterns) ra file Excel."""
    if request.method != 'POST':
        return JsonResponse({"status": "error", "message": "Invalid method"}, status=405)

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from io import BytesIO
        import numpy as np

        data = json.loads(request.body)

        # Lấy lại thông số từ request
        length = int(data['length'])
        te_dau_sat = int(data['te_dau_sat'])
        blade_width = float(data['blade_width'])
        hao_hut_percent = float(data.get('hao_hut_percent', 1))
        pieces_data = data.get('pieces_data', [])
        valid_pieces = [row for row in pieces_data if row and row[0] is not None and row[1] is not None and row[2] is not None]

        if not valid_pieces:
            return JsonResponse({"status": "error", "message": "Không có dữ liệu đoạn cắt."}, status=400)

        piece_names = [str(item[0]) for item in valid_pieces]
        segment_sizes = [float(item[1]) for item in valid_pieces]
        demands = [int(item[2]) for item in valid_pieces]

        # Load patterns từ cache (giống logic optimize_cutting)
        from .optimization_logic import SteelCuttingOptimizer
        optimizer = SteelCuttingOptimizer(
            length=length,
            te_dau_sat=te_dau_sat,
            piece_names=piece_names,
            segment_sizes=segment_sizes,
            demands=demands,
            blade_width=blade_width,
            factors=[1],
            max_manual_cuts=0,
            max_stock_over=0,
            hao_hut_percent=hao_hut_percent,
            time_limit_seconds=30
        )

        solutions = optimizer.load_solution_from_pickle()
        if not solutions:
            return JsonResponse({"status": "error", "message": "Chưa có dữ liệu GĐ 1. Hãy chạy tối ưu trước."}, status=404)

        # --- Tạo Excel workbook ---
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Kết quả GĐ1"

        # Styles
        header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Tiêu đề phụ (thông số)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(piece_names) + 4)
        title_cell = ws.cell(row=1, column=1,
                             value=f"DANH SÁCH PATTERN GĐ 1 — Sắt {length}mm | Tề đầu {te_dau_sat}mm | Lưỡi {blade_width}mm | Hao hụt cho phép {hao_hut_percent}%")
        title_cell.font = Font(name='Arial', bold=True, size=13, color='1F4E79')
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 30

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(piece_names) + 4)
        info_cell = ws.cell(row=2, column=1,
                            value=f"Tổng số pattern: {len(solutions)} | Đoạn cắt: {', '.join(piece_names)}")
        info_cell.font = Font(name='Arial', size=10, italic=True, color='555555')
        info_cell.alignment = Alignment(horizontal='left')

        # Headers (dòng 4)
        header_row = 4
        headers = ['STT']
        for i, name in enumerate(piece_names):
            size_str = f"{int(segment_sizes[i])}" if segment_sizes[i] == int(segment_sizes[i]) else f"{segment_sizes[i]:.1f}"
            headers.append(f"{name}\n({size_str}mm)")
        headers.extend(['Tổng sắt dùng (mm)', 'Hao hụt (mm)', 'Hao hụt (%)'])

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        ws.row_dimensions[header_row].height = 35

        # Data rows
        for row_idx, (obj_value, pattern) in enumerate(solutions, 1):
            data_row = header_row + row_idx
            waste = length - obj_value - te_dau_sat
            waste_pct = (waste / length * 100) if length > 0 else 0

            values = [row_idx]
            for i in range(len(piece_names)):
                values.append(pattern[i] if i < len(pattern) else 0)
            values.extend([round(obj_value, 1), round(waste, 1), round(waste_pct, 2)])

            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=data_row, column=col_idx, value=val)
                cell.alignment = cell_align
                cell.border = thin_border

                # Highlight pattern quantities > 0
                if 1 < col_idx <= len(piece_names) + 1 and val > 0:
                    cell.font = Font(name='Arial', bold=True, color='1F4E79')

        # Column widths
        ws.column_dimensions['A'].width = 6
        for col_idx in range(2, len(piece_names) + 2):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 16
        ws.column_dimensions[openpyxl.utils.get_column_letter(len(piece_names) + 2)].width = 20
        ws.column_dimensions[openpyxl.utils.get_column_letter(len(piece_names) + 3)].width = 15
        ws.column_dimensions[openpyxl.utils.get_column_letter(len(piece_names) + 4)].width = 13

        # Auto filter
        last_col_letter = openpyxl.utils.get_column_letter(len(headers))
        last_data_row = header_row + len(solutions)
        ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{last_data_row}"

        # Freeze panes
        ws.freeze_panes = f"A{header_row + 1}"

        # Tạo sheet tóm tắt nhu cầu
        ws2 = wb.create_sheet(title="Nhu cầu cắt")
        ws2.cell(row=1, column=1, value="Tên sắt").font = Font(bold=True)
        ws2.cell(row=1, column=2, value="Kích thước (mm)").font = Font(bold=True)
        ws2.cell(row=1, column=3, value="SL Cần").font = Font(bold=True)
        for i, (name, size, demand) in enumerate(zip(piece_names, segment_sizes, demands)):
            ws2.cell(row=i + 2, column=1, value=name)
            ws2.cell(row=i + 2, column=2, value=size)
            ws2.cell(row=i + 2, column=3, value=demand)
        ws2.column_dimensions['A'].width = 15
        ws2.column_dimensions['B'].width = 18
        ws2.column_dimensions['C'].width = 12

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        response['Content-Disposition'] = f'attachment; filename="KetQua_GD1_Sat{length}mm_{timestamp}.xlsx"'
        return response

    except Exception as e:
        logger.error(f"Export Excel Phase 1 failed: {traceback.format_exc()}")
        return JsonResponse({"status": "error", "message": str(e)}, status=500)